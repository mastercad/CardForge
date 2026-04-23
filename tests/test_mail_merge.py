"""Tests für cardforge.mail_merge – Datei-Lade- und Merge-Logik sowie Dialog."""

from __future__ import annotations

from cardforge.mail_merge import _apply_merge, _load_csv, _load_excel
from cardforge.models import ELEMENT_TEXT, CardElement, CardLayout

# ---------------------------------------------------------------------------
# _load_csv
# ---------------------------------------------------------------------------


class TestLoadCsv:
    def test_basic_csv(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("Name,Email\nAlice,alice@example.com\nBob,bob@example.com", encoding="utf-8")
        headers, rows = _load_csv(str(f))
        assert headers == ["Name", "Email"]
        assert len(rows) == 2
        assert rows[0]["Name"] == "Alice"
        assert rows[1]["Email"] == "bob@example.com"

    def test_utf8_bom(self, tmp_path):
        f = tmp_path / "bom.csv"
        f.write_bytes(b"\xef\xbb\xbfSpalte\nWert\n")
        headers, rows = _load_csv(str(f))
        assert "Spalte" in headers

    def test_empty_csv(self, tmp_path):
        f = tmp_path / "empty.csv"
        f.write_text("Name,Email\n", encoding="utf-8")
        headers, rows = _load_csv(str(f))
        assert headers == ["Name", "Email"]
        assert rows == []

    def test_csv_no_header(self, tmp_path):
        f = tmp_path / "noheader.csv"
        f.write_text("", encoding="utf-8")
        headers, rows = _load_csv(str(f))
        assert rows == []

    def test_semicolon_delimiter_not_supported_gracefully(self, tmp_path):
        """Semikolon-Delimiter ergibt eine einzige Spalte (kein crash)."""
        f = tmp_path / "semi.csv"
        f.write_text("Name;Email\nAlice;alice@test.com\n", encoding="utf-8")
        headers, rows = _load_csv(str(f))
        # Kein crash, Anzahl Zeilen stimmt
        assert isinstance(rows, list)


# ---------------------------------------------------------------------------
# _load_excel
# ---------------------------------------------------------------------------


class TestLoadExcel:
    def _make_xlsx(self, tmp_path, headers, rows):
        """Erstellt eine temporäre XLSX-Datei."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        path = str(tmp_path / "data.xlsx")
        wb.save(path)
        return path

    def test_basic_excel(self, tmp_path):
        path = self._make_xlsx(tmp_path, ["Name", "Firma"], [["Alice", "ACME"], ["Bob", "Corp"]])
        headers, rows = _load_excel(path)
        assert headers == ["Name", "Firma"]
        assert len(rows) == 2
        assert rows[0]["Name"] == "Alice"
        assert rows[1]["Firma"] == "Corp"

    def test_empty_excel(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        path = str(tmp_path / "empty.xlsx")
        wb.save(path)
        headers, rows = _load_excel(path)
        assert headers == []
        assert rows == []

    def test_none_cells_become_empty_string(self, tmp_path):
        path = self._make_xlsx(tmp_path, ["Name", "Opt"], [["Alice", None]])
        _, rows = _load_excel(path)
        assert rows[0]["Opt"] == ""

    def test_numeric_values_stringified(self, tmp_path):
        path = self._make_xlsx(tmp_path, ["Zahl"], [[42]])
        _, rows = _load_excel(path)
        assert rows[0]["Zahl"] == "42"


# ---------------------------------------------------------------------------
# _apply_merge
# ---------------------------------------------------------------------------


class TestApplyMerge:
    def _layout_with_template(self, text: str) -> CardLayout:
        layout = CardLayout()
        layout.front_elements.append(CardElement(type=ELEMENT_TEXT, text=text))
        return layout

    def test_single_placeholder(self):
        layout = self._layout_with_template("Hallo {{Name}}")
        merged = _apply_merge(layout, {"Name": "Alice"})
        assert merged.front_elements[0].text == "Hallo Alice"

    def test_multiple_placeholders(self):
        layout = self._layout_with_template("{{Vorname}} {{Nachname}}")
        merged = _apply_merge(layout, {"Vorname": "Max", "Nachname": "Muster"})
        assert merged.front_elements[0].text == "Max Muster"

    def test_unknown_placeholder_unchanged(self):
        layout = self._layout_with_template("{{Unbekannt}}")
        merged = _apply_merge(layout, {"Name": "Alice"})
        assert merged.front_elements[0].text == "{{Unbekannt}}"

    def test_does_not_mutate_original(self):
        layout = self._layout_with_template("{{Name}}")
        original_text = layout.front_elements[0].text
        _apply_merge(layout, {"Name": "Alice"})
        assert layout.front_elements[0].text == original_text

    def test_back_elements_also_merged(self):
        layout = CardLayout()
        layout.back_elements.append(CardElement(type=ELEMENT_TEXT, text="{{Firma}}"))
        merged = _apply_merge(layout, {"Firma": "ACME"})
        assert merged.back_elements[0].text == "ACME"

    def test_non_text_elements_not_affected(self):
        from cardforge.models import ELEMENT_RECT

        layout = CardLayout()
        layout.front_elements.append(CardElement(type=ELEMENT_RECT, text="{{Name}}"))
        # ELEMENT_RECT wird nicht gemergt (nur ELEMENT_TEXT)
        merged = _apply_merge(layout, {"Name": "Alice"})
        assert merged.front_elements[0].text == "{{Name}}"

    def test_empty_row(self):
        layout = self._layout_with_template("{{Name}}")
        merged = _apply_merge(layout, {})
        assert merged.front_elements[0].text == "{{Name}}"

    def test_multiple_occurrences_replaced(self):
        layout = self._layout_with_template("{{Name}} und nochmal {{Name}}")
        merged = _apply_merge(layout, {"Name": "Bob"})
        assert merged.front_elements[0].text == "Bob und nochmal Bob"


# ---------------------------------------------------------------------------
# MailMergeDialog
# ---------------------------------------------------------------------------


class TestMailMergeDialog:
    def test_creates_without_crash(self, qapp):
        from cardforge.mail_merge import MailMergeDialog

        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        assert dlg is not None
        dlg.close()

    def test_initial_state(self, qapp):
        from cardforge.mail_merge import MailMergeDialog

        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        assert dlg._headers == []  # noqa: SLF001
        assert dlg._rows == []  # noqa: SLF001
        dlg.close()

    def test_result_layouts_initially_empty(self, qapp):
        from cardforge.mail_merge import MailMergeDialog

        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        assert dlg._result_layouts == []  # noqa: SLF001
        dlg.close()

    def test_load_file_csv_populates_table(self, qapp, tmp_path):
        from unittest.mock import patch

        from cardforge.mail_merge import MailMergeDialog

        csv_file = tmp_path / "data.csv"
        csv_file.write_text("Name,Email\nAlice,alice@test.com\n", encoding="utf-8")
        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        with patch(
            "cardforge.mail_merge.QFileDialog.getOpenFileName", return_value=(str(csv_file), "")
        ):
            dlg._load_file()  # noqa: SLF001
        assert dlg._headers == ["Name", "Email"]  # noqa: SLF001
        assert len(dlg._rows) == 1  # noqa: SLF001
        assert dlg._table.columnCount() == 2  # noqa: SLF001
        assert dlg._ph_list.count() == 2  # noqa: SLF001
        dlg.close()

    def test_load_file_excel_populates_table(self, qapp, tmp_path):
        from unittest.mock import patch

        import openpyxl

        from cardforge.mail_merge import MailMergeDialog

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Firma", "Stadt"])
        ws.append(["ACME", "Berlin"])
        xls_path = str(tmp_path / "data.xlsx")
        wb.save(xls_path)

        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        with patch("cardforge.mail_merge.QFileDialog.getOpenFileName", return_value=(xls_path, "")):
            dlg._load_file()  # noqa: SLF001
        assert dlg._headers == ["Firma", "Stadt"]  # noqa: SLF001
        dlg.close()

    def test_load_file_cancel_noop(self, qapp):
        from unittest.mock import patch

        from cardforge.mail_merge import MailMergeDialog

        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        with patch("cardforge.mail_merge.QFileDialog.getOpenFileName", return_value=("", "")):
            dlg._load_file()  # noqa: SLF001
        assert dlg._headers == []  # noqa: SLF001
        dlg.close()

    def test_load_file_error_shows_dialog(self, qapp, tmp_path):
        from unittest.mock import patch

        from cardforge.mail_merge import MailMergeDialog

        csv_file = tmp_path / "bad.csv"
        csv_file.write_bytes(b"\xff\xfe broken content")
        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        with (
            patch(
                "cardforge.mail_merge.QFileDialog.getOpenFileName", return_value=(str(csv_file), "")
            ),
            patch("cardforge.mail_merge._load_csv", side_effect=RuntimeError("oops")),
            patch("cardforge.mail_merge.QMessageBox.critical") as mock_crit,
        ):
            dlg._load_file()  # noqa: SLF001
            mock_crit.assert_called_once()
        dlg.close()

    def test_generate_with_rows_accepts(self, qapp, tmp_path):
        from unittest.mock import patch

        from cardforge.mail_merge import MailMergeDialog

        csv_file = tmp_path / "data.csv"
        csv_file.write_text("Name\nAlice\nBob\n", encoding="utf-8")
        layout = CardLayout(name="Template")
        layout.front_elements.append(CardElement(type=ELEMENT_TEXT, text="{{Name}}"))
        dlg = MailMergeDialog(layout)
        with patch(
            "cardforge.mail_merge.QFileDialog.getOpenFileName", return_value=(str(csv_file), "")
        ):
            dlg._load_file()  # noqa: SLF001
        # now generate
        with patch.object(dlg, "accept") as mock_accept:
            dlg._generate()  # noqa: SLF001
            mock_accept.assert_called_once()
        assert len(dlg._result_layouts) == 2  # noqa: SLF001
        assert dlg._result_layouts[0].name == "Template [1]"  # noqa: SLF001

    def test_generate_with_empty_rows_warns(self, qapp):
        from unittest.mock import patch

        from cardforge.mail_merge import MailMergeDialog

        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        with patch("cardforge.mail_merge.QMessageBox.warning") as mock_warn:
            dlg._generate()  # noqa: SLF001
            mock_warn.assert_called_once()
        assert dlg._result_layouts == []  # noqa: SLF001
        dlg.close()

    def test_populate_table_more_than_50_rows(self, qapp):
        from cardforge.mail_merge import MailMergeDialog

        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        dlg._headers = ["Name"]  # noqa: SLF001
        dlg._rows = [{"Name": f"Person {i}"} for i in range(60)]  # noqa: SLF001
        dlg._populate_table()  # noqa: SLF001
        assert dlg._table.rowCount() == 50  # capped at 50
        dlg.close()

    def test_result_layouts_accessor(self, qapp):
        from cardforge.mail_merge import MailMergeDialog

        layout = CardLayout(name="Template")
        dlg = MailMergeDialog(layout)
        dlg._result_layouts = [CardLayout(name="X")]  # noqa: SLF001
        result = dlg.result_layouts()
        assert len(result) == 1
        assert result[0].name == "X"
        dlg.close()
